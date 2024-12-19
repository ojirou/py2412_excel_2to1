import openpyxl
import subprocess
from copy import copy
input_filename = r"sample_2to1\\sample.xlsx"
output_filename = r"sample_2to1\\sample_output.xlsx"
# 新しいワークブックを作成
wb = openpyxl.Workbook()
ws = wb.active
ws2 = wb.create_sheet("Sheet1")
# 元のExcelファイルを読み込む
source_wb = openpyxl.load_workbook(input_filename)
source_ws = source_wb["Sheet1"]
# 行を1から末尾まで処理
dest_row=1
for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, min_col=1, max_col=2):
    cell_A = row[0]  # A列の値
    cell_B = row[1]  # B列の値
    if cell_A.value is not None:
        ws[f"A{dest_row}"].value=cell_A.value
        if cell_A.hyperlink:
            ws[f"A{dest_row}"].hyperlink=cell_A.hyperlink.target
        ws[f"A{dest_row}"].font=copy(cell_A.font)
        ws[f"A{dest_row}"].fill=copy(cell_A.fill)
        ws[f"A{dest_row}"].border=copy(cell_A.border)
        ws[f"A{dest_row}"].alignment=copy(cell_A.alignment)
        ws[f"A{dest_row}"].number_format=copy(cell_A.number_format)
        ws[f"A{dest_row}"].protection=copy(cell_A.protection)
        dest_row+=1
    if cell_B.value is not None:
        ws[f"A{dest_row}"].value=cell_B.value
        if cell_B.hyperlink:
           ws[f"A{dest_row}"].hyperlink=cell_B.hyperlink.target
        ws[f"A{dest_row}"].font=copy(cell_B.font)
        ws[f"A{dest_row}"].fill=copy(cell_B.fill)
        ws[f"A{dest_row}"].border=copy(cell_B.border)
        ws[f"A{dest_row}"].alignment=copy(cell_B.alignment)
        ws[f"A{dest_row}"].number_format=copy(cell_B.number_format)
        ws[f"A{dest_row}"].protection=copy(cell_B.protection)
        dest_row+=1          
# 保存
wb.save(output_filename)
# 出力ファイルを開く
subprocess.Popen(["start", "", output_filename], shell=True)